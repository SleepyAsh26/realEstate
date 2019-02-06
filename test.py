#PLOT CORELATION HEATMAP

fig, ax = plt.subplots()
ax.bar(y_rbf, y_test)
ax.set(xlabel="Real", ylabel="Prediction");
plt.show()

# Correction Matrix Plot
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import seaborn as sns
 
df = pd.read_excel("SSandMH_clear_gps.xlsx")
df = df.drop(["ID", "PerSq", "Price", 
              "Street", "Gas", "Storeroom", 
              "Floors","Parking", "Lat", "Lng", 
              "District", "Condition", "Cluster",
              "Area", "Rooms", "Bedrooms"], axis=1)
#X = pd.get_dummies(X, columns=["Status", "Condition", "District"], prefix=["St", "Co", "Di"], drop_first=True)
df = pd.get_dummies(df, columns=["Status"], prefix=["St"])#, drop_first=True


A = df["Status"]
A = pd.get_dummies(A, columns=["Status"], prefix=["St"], drop_first=True)
B = df["Heating"]
B = B.to_frame()

A = list(A.values)
B = list(B.values)

lst = []
for i in range(len(A)):
    if A[i] == 0 and B[i] == 1:
        lst.append(1)
    else:
        lst.append(0)
print("{}".format(lst.count(1)))
print("{}".format(lst.count(0)))


corr = df.corr()
#Plot figsize
fig, ax = plt.subplots()
#Generate Color Map, red & blue
colormap = sns.diverging_palette(220, 30, as_cmap=True)
#Generate Heat Map, allow annotations and place floats in map
sns.heatmap(corr, cmap=colormap, annot=True, fmt=".2f")
#Apply xticks
plt.xticks(range(len(corr.columns)), corr.columns);
#Apply yticks
plt.yticks(range(len(corr.columns)), corr.columns)
#show plot
plt.show()



#GEO TO ENG FOR COLUMS
GeToEn = {'ა':'a',
          'ბ':'b',
          'გ':'g',
          'დ':'d',
          'ე':'e',
          'ვ':'v',
          'ზ':'z',
          'თ':'t',
          'ი':'i',
          'კ':'k',
          'ლ':'l',
          'მ':'m',
          'ნ':'n',
          'ო':'o',
          'პ':'p',
          'ჟ':'dj',
          'რ':'r',
          'ს':'s',
          'ტ':'t',
          'უ':'u',
          'ფ':'f',
          'ქ':'q',
          'ღ':'gh',
          'ყ':'kh',
          'შ':'sh',
          'ჩ':'ch',
          'ც':'c',
          'ძ':'zh',
          'წ':'ts',
          'ჭ':'ts',
          'ხ':'kh',
          'ჯ':'j',
          'ჰ':'h',
          ' ':'_',
          ',':'/',
          '#':'No',
          '(':'',
          ')':'',
          '/':'',
          '\\':''}

# Importing the dataset
from openpyxl import load_workbook
workbook = load_workbook(filename = "SSandMH.xlsx")
ws = workbook.active
row = ws.max_row
col = 4
for i in range(2,row+1):
    try:
        string = ws.cell(row=i, column=col).value
        for char in string:
            try:
                string = string.replace(char, GeToEn[char])
            except:
                pass
        ws.cell(row=i, column=col).value = string
    except:
        pass
workbook.save('SSandMH.xlsx')

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb  5 06:36:29 2019

@author: sleepyash
"""

from sklearn.svm import SVR
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import train_test_split
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
import matplotlib
import pandas as pd
import numpy as np

from sklearn.cluster import DBSCAN
from geopy.distance import great_circle
from shapely.geometry import MultiPoint

from sklearn.cluster import KMeans
from collections import Counter

df = pd.read_excel("SSandMH_clear_gps.xlsx")
coords = df.as_matrix(columns=['Lat', 'Lng'])

df1 = df[df["Cluster"] != -1]

  
import seaborn as sns
sns.lmplot(x='Lng', y='Lat', data=df, hue='Cluster', fit_reg=False)#palette="tab20c",
plt.show()

"""
Kmean = KMeans(n_clusters=100).fit(coords)
K = Kmean.cluster_centers_

import seaborn as sns
sns.lmplot('Lng', 'Lat', data=df, hue='Cluster', fit_reg=False)
plt.show()

#Kmean.labels_
P = Kmean.predict(coords)

from openpyxl import load_workbook
workbook = load_workbook(filename = "SSandMH_clear_gps.xlsx")
ws = workbook.active
for i, elem in enumerate(P):
    print(f"{i}/{len(P)}")
    ws.cell(row=i+2, column=8).value = elem
workbook.save('SSandMH_clear_gps.xlsx')
"""

"""
lats, lons = zip(*K)
rep_points = pd.DataFrame({'Lng':lons, 'Lat':lats})
rs = rep_points#.apply(lambda row: df[(df['Lat']==row['Lat']) & (df['Lng']==row['Lng'])].iloc[0], axis=1)

fig, ax = plt.subplots(figsize=[10, 6])
rs_scatter = ax.scatter(rs['Lng'], rs['Lat'], c='#99cc99', edgecolor='None', alpha=0.7, s=120)
df_scatter = ax.scatter(df['Lng'], df['Lat'], c='k', alpha=0.9, s=3)
ax.set_title('Full data set vs DBSCAN reduced set')
ax.set_xlabel('Longitude')
ax.set_ylabel('Latitude')
ax.legend([df_scatter, rs_scatter], ['Full set', 'Reduced set'], loc='upper right')
plt.show()
"""


kms_per_radian = 6371.0088
epsilon = 0.5/kms_per_radian # 1.5km
                        #min_samples 1 NO noise everything metters
db = DBSCAN(eps=epsilon, min_samples=5, algorithm='ball_tree', metric='haversine').fit(np.radians(coords))
cluster_labels = db.labels_
num_clusters = len(set(cluster_labels))
clusters = pd.Series([coords[cluster_labels == n] for n in range(num_clusters)])
print('Number of clusters: {}'.format(num_clusters))


from openpyxl import load_workbook
workbook = load_workbook(filename = "SSandMH_clear_gps.xlsx")
ws = workbook.active
for i, elem in enumerate(cluster_labels):
    print(f"{i}/{len(cluster_labels)}")
    ws.cell(row=i+2, column=8).value = elem
workbook.save('SSandMH_clear_gps.xlsx')


"""
def get_centermost_point(cluster):
    centroid = (MultiPoint(cluster).centroid.x, MultiPoint(cluster).centroid.y)
    centermost_point = min(cluster, key=lambda point: great_circle(point, centroid).m)
    return tuple(centermost_point)
centermost_points = clusters.map(get_centermost_point)

lats, lons = zip(*centermost_points)
rep_points = pd.DataFrame({'Lng':lons, 'Lat':lats})

rs = rep_points.apply(lambda row: df[(df['Lat']==row['Lat']) & (df['Lng']==row['Lng'])].iloc[0], axis=1)

fig, ax = plt.subplots(figsize=[10, 6])
rs_scatter = ax.scatter(rs['Lng'], rs['Lat'], c='#99cc99', edgecolor='None', alpha=0.7, s=120)
df_scatter = ax.scatter(df['Lng'], df['Lat'], c='k', alpha=0.9, s=3)
ax.set_title('Full data set vs DBSCAN reduced set')
ax.set_xlabel('Longitude')
ax.set_ylabel('Latitude')
ax.legend([df_scatter, rs_scatter], ['Full set', 'Reduced set'], loc='upper right')
plt.show()
"""



















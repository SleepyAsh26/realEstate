def scramble(s1, s2):
    s3 = ''
    for i in s2:
        if i in s1:
            s3 += i
            s1 = list(s1)
            s1 = s1.remove(i)
            s1 = str(s1)
    print(s1,s2,s3)
    return True if s3 == s2 else False

print(scramble("tjuofkoiyufmdgfhn", "fodjhfuyuf"))
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd




GeToEn = {'ა':'a',
          'ბ':'b',
          'გ':'g',
          'დ':'d',
          'ე':'e',
          'ვ':'v',
          'ზ':'z',
          'თ':'t',
          'ი':'i',
          'კ':'k',
          'ლ':'l',
          'მ':'m',
          'ნ':'n',
          'ო':'o',
          'პ':'p',
          'ჟ':'dj',
          'რ':'r',
          'ს':'s',
          'ტ':'t',
          'უ':'u',
          'ფ':'f',
          'ქ':'q',
          'ღ':'gh',
          'ყ':'kh',
          'შ':'sh',
          'ჩ':'ch',
          'ც':'c',
          'ძ':'zh',
          'წ':'ts',
          'ჭ':'ts',
          'ხ':'kh',
          'ჯ':'j',
          'ჰ':'h',
          ' ':'_',
          ',':'/',
          '#':'No',
          '(':'',
          ')':'',
          '/':'',
          '\\':''}

# Importing the dataset
from openpyxl import load_workbook
workbook = load_workbook(filename = "SSandMH.xlsx")
ws = workbook.active
row = ws.max_row
col = 4
for i in range(2,row+1):
    try:
        string = ws.cell(row=i, column=col).value
        for char in string:
            try:
                string = string.replace(char, GeToEn[char])
            except:
                pass
        ws.cell(row=i, column=col).value = string
    except:
        pass
workbook.save('SSandMH.xlsx')

from openpyxl import load_workbook
import pandas as pd

workbook = load_workbook(filename = "Map.xlsx")
ws = workbook.active
dataset = pd.read_excel('Map.xlsx')
data = dataset["Address"]
for i,j in enumerate(data):
    lat, lng = j.split(', ')
    ws.cell(row=i+2, column=5).value = lat
    ws.cell(row=i+2, column=6).value = lng
workbook.save('Map.xlsx')


workbook = load_workbook(filename = "saburtalo.xlsx")
ws = workbook.active
col = 6
add = []
for i in range(2,831):
    add.append(ws.cell(row=i, column=col).value)
rem = []
for j in add:
    if add.count(j) > 5:
        pass
    else:
        rem.append(j)
for i in range(1,831):
    string = ws.cell(row=i, column=col).value
    if string in rem:
        ws.cell(row=i, column=col).value = 'X'
workbook.save('saburtalo.xlsx')




import timeit

def scramble():
    s1 = "tjuofkoiyufmdgfhn"
    s2 = "fodjhfuyuf"
    for i in s2:
        if i in s1:
            s1 = s1.replace(i, '', 1)
        else:
            return False
    
    return True
# 2.2785653109967825
print( timeit.timeit(scramble, number=1000000) )

a=[1,2,3,4,6]
b=[1,3,5,2,2,6]
res = [char for char in b if char in a]
res2 = (char for char in b if char in a)
print(res, res2)

a = "Hello, World!"
print(a.lower())
print(a)


def scramble():
    s1 = "tjuofkoiyufmdgfhn"
    s2 = "fodjhfuyuf"
    for i in s1:
        s2 = s2.replace(i,'', 1)
    return s2 == ''

def scramble():
    s1 = "tjuofkoiyufmdgfhn"
    s2 = "fodjhfuyuf"
    s1 = list(s1)
    for i in s2:
        try:
            s1.remove(i)
        except:
            return False
    return True

def scramble():
    s1 = "tjuofkoiyufmdgfhn"
    s2 = "fodjhfuyuf"
    for i in s2:
        if s1.count(i) >= s2.count(i):
            return False
    return True

def scramble(s1, s2):
    
    res = []
    for char in s2:
        if char in s1:
            res.append(char)
            s1 = s1.replace(char, "", 1)
    #res = ''.join([char for char in s2 if char in s1])
    print(s1)
    print(s2)
    print(res)
    res = ''.join(res)
    x = res == s2
    print(x)
    return x
a = [1,2,3,4,5,6,7,8,9]
print(a[0:9:3])
    
def primeFactors(n):
    rem = n
    res = ''
    for prime in range(2,n):
        if any(prime % num == 0 for num in range(2, prime)):
            continue
        else:
            print(prime, rem)
            for power in range(0,10):
                if rem % prime == 0:
                    rem = rem/prime
                else:
                    break
                print(rem)
                
            res += "{}{}{}".format('('+str(prime) if power != 0 else '' ,
                                    "**"+str(power) if power > 1 else '',
                                    ")" if power != 0 else '')
            
            if rem == 1:
                #res += "("+str(int(rem))+")" if rem != 1 else ''
                break
    return res

print(primeFactors(18195729))


print(sum(list(99)))

for possiblePrime in range(2,21):
    isPrime = True
    for num in range(2, possiblePrime):
        if possiblePrime % num == 0:
            isPrime = False
    if isPrime:
        pass
        """
        i = possiblePrime
        power = 0
        while rem % i == 0:
            rem = rem / i
            power += 1
        if power != 0:
            if power > 1:
                res = res + "("+str(i)+"**"+str(power)+")"
            else:
                res = res + "("+str(i)+")"
        if rem == 1:
            break
            """
return res
            
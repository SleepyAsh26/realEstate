#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jan  4 04:14:10 2019

@author: sleepyash
"""
#AIzaSyDhw9EWzXaeWfpi1fMfZiRR7Y6tS-fKIUw 
import googlemaps
import pandas as pd
import gmaps
from ipywidgets.embed import embed_minimal_html
from shapely.geometry import Point, Polygon
import json
from openpyxl import load_workbook

api_key = 'AIzaSyB47j8P2l_IEm_8mg-sUltopgKIya_PMt8'

Gmaps = googlemaps.Client(key=api_key)
dataset = pd.read_excel('SSandMH_clear_gps.xlsx')
df = df[df["Cluster"] != -1]
coords = df.as_matrix(columns=['Lat', 'Lng'])

data = dataset["Street"]
data = list(data)
"""
Lats = dataset["Lat"]
Lngs = dataset["Lng"]
coors = []
for i in range(len(Lats)):
    coors.append([Lats[i],Lngs[i]])
"""
# Geocoding an address
coordinates = []
TbilisiPLY = [(41.816420, 44.680124),
              (41.831395, 44.876235),
              (41.772717, 44.977507),
              (41.673009, 44.999999),#45.007487)
              (41.656818, 44.627870)]
poly = Polygon(TbilisiPLY)




workbook = load_workbook(filename = "SSandMH_clear_gps.xlsx")
ws = workbook.active
for i, place in enumerate(data):
    print(f"{i}/{len(data)}")
    if place != 'nan':
        try:
            geocode_result = Gmaps.geocode(place)
            coor = geocode_result[0]
            coor = coor["geometry"]
            coor = coor['location']
            lat = round(coor['lat'],7)
            lng = round(coor['lng'],7)
            if poly.contains(Point(lat,lng)) == True:
                coordinates.append([lat,lng])
                ws.cell(row=i+2, column=6).value = lat
                ws.cell(row=i+2, column=7).value = lng
                print(lat,lng)
            else:
                coordinates.append("")
                print("Nan")
        except:
            pass
workbook.save('SSandMH_clear_gps.xlsx')
        






tbilisi_file = open('polygons/tbilisi.json')
tbilisi_str = tbilisi_file.read()
tbilisi_ply = json.loads(tbilisi_str)['coordinates'][0][0]
for i in range(len(tbilisi_ply)):
    temp = tbilisi_ply[i][0]
    tbilisi_ply[i][0] = tbilisi_ply[i][1]
    tbilisi_ply[i][1] = temp

DCh_file = open('polygons/Didube-Chugureti.json')
DCh_str = DCh_file.read()
DCh_ply = json.loads(DCh_str)['geometries'][0]
DCh_ply = DCh_ply['coordinates'][0]
DCh_ply = DCh_ply[0]
for i in range(len(DCh_ply)):
    temp = DCh_ply[i][0]
    DCh_ply[i][0] = DCh_ply[i][1]
    DCh_ply[i][1] = temp

tbilisi_file = open('polygons/Gldani-Nadzaladevi.json')
tbilisi_str = tbilisi_file.read()
tbilisi_ply = json.loads(tbilisi_str)['coordinates'][0][0]
for i in range(len(tbilisi_ply)):
    temp = tbilisi_ply[i][0]
    tbilisi_ply[i][0] = tbilisi_ply[i][1]
    tbilisi_ply[i][1] = temp

tbilisi_file = open('polygons/Isani-Samgori.json')
tbilisi_str = tbilisi_file.read()
tbilisi_ply = json.loads(tbilisi_str)['coordinates'][0][0]
for i in range(len(tbilisi_ply)):
    temp = tbilisi_ply[i][0]
    tbilisi_ply[i][0] = tbilisi_ply[i][1]
    tbilisi_ply[i][1] = temp
    
tbilisi_file = open('polygons/Old-Tbilisi.json')
tbilisi_str = tbilisi_file.read()
tbilisi_ply = json.loads(tbilisi_str)['coordinates'][0][0]
for i in range(len(tbilisi_ply)):
    temp = tbilisi_ply[i][0]
    tbilisi_ply[i][0] = tbilisi_ply[i][1]
    tbilisi_ply[i][1] = temp    

tbilisi_file = open('polygons/Vake-Saburtalo.json')
tbilisi_str = tbilisi_file.read()
tbilisi_ply = json.loads(tbilisi_str)['coordinates'][0][0]
for i in range(len(tbilisi_ply)):
    temp = tbilisi_ply[i][0]
    tbilisi_ply[i][0] = tbilisi_ply[i][1]
    tbilisi_ply[i][1] = temp    
    
    
    
    
saburtalo_file = open('saburtalo.json')
saburtalo_str = saburtalo_file.read()
saburtalo_ply = json.loads(saburtalo_str)['coordinates'][0]
for i in range(len(saburtalo_ply)):
    temp = saburtalo_ply[i][0]
    saburtalo_ply[i][0] = saburtalo_ply[i][1]
    saburtalo_ply[i][1] = temp

gmaps.configure(api_key=api_key)

TbilisiPLY = [(41.816420, 44.680124),
              (41.831395, 44.876235),
              (41.772717, 44.977507),
              (41.673009, 45.007487),
              (41.656818, 44.627870)]

DidiDigomiPLY = [(41.801125, 44.726993), 
                 (41.821502, 44.776487),
                 (41.814003, 44.784271), 
                 (41.787966, 44.784622),
                 (41.787665, 44.770549), 
                 (41.760436, 44.768482),
                 (41.766687, 44.709715)]

DidiDigomi = []
poly = Polygon(TbilisiPLY)
for Clat, Clng in coordinates:
    if poly.contains(Point(Clat,Clng)) == True:
        DidiDigomi.append([Clat,Clng])
for i in enumerate():

gmaps.configure(api_key=api_key)
 
fig = gmaps.figure()

fig.add_layer(gmaps.drawing_layer(features=[gmaps.Polygon(tbilisi_ply, stroke_color='red', fill_color=(255, 0, 132)) ]))
fig.add_layer(gmaps.drawing_layer(features=[gmaps.Polygon(saburtalo_ply, stroke_color='blue', fill_color=(255, 0, 132)) ]))

heatmap_layer = gmaps.heatmap_layer(coords)
fig.add_layer(heatmap_layer)

markers = gmaps.symbol_layer(coors, fill_color='green', stroke_color='green', scale=2)
fig.add_layer(markers)
embed_minimal_html('export.html', views=[fig])